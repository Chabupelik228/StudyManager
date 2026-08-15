import asyncio
import calendar
from contextlib import suppress
from datetime import datetime
import html
import io
import json
import logging
import os
import re
import time
import uuid

import aiohttp
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Command, CommandStart
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaDocument,
    InputMediaPhoto,
    Message,
)
import asyncpg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pytz

load_dotenv()

# --- CONFIG & SECRETS ---
TOKEN = os.getenv("BOT_TOKEN")
GROUP_ID = int(os.getenv("GROUP_ID", 0))
VK_TOKEN = os.getenv("VK_TOKEN")
VK_GROUP_ID = int(os.getenv("VK_GROUP_ID", 0))
VK_CHAT_PEER_ID = int(os.getenv("VK_CHAT_PEER_ID", 0))
VK_API_VERSION = "5.199"

TEST_GROUP_ID = int(os.getenv("TEST_GROUP_ID", 0))
TEST_VK_CHAT_PEER_ID = int(os.getenv("TEST_VK_CHAT_PEER_ID", 0))

CHAT_MAP = {
    VK_CHAT_PEER_ID: GROUP_ID,
}
if TEST_VK_CHAT_PEER_ID and TEST_GROUP_ID:
    CHAT_MAP[TEST_VK_CHAT_PEER_ID] = TEST_GROUP_ID

REVERSE_CHAT_MAP = {v: k for k, v in CHAT_MAP.items() if v != 0 and k != 0}
PROCESSED_VK_MSGS = set()

# PostgreSQL settings
POSTGRES_USER = os.getenv("POSTGRES_USER", "studymanager")
POSTGRES_PASSWORD = os.getenv("POSTGRES_PASSWORD", "StudyManager2024!Secure")
POSTGRES_DB = os.getenv("POSTGRES_DB", "studymanager_db")
POSTGRES_HOST = os.getenv("POSTGRES_HOST", "db")
POSTGRES_PORT = int(os.getenv("POSTGRES_PORT", 5432))
BACKEND_URL = os.getenv("BACKEND_URL", "http://backend:8000")

# Database Pool
db_pool: asyncpg.Pool | None = None


async def get_db_pool() -> asyncpg.Pool:
    global db_pool
    if db_pool is None or db_pool._closed:
        db_pool = await asyncpg.create_pool(
            user=POSTGRES_USER,
            password=POSTGRES_PASSWORD,
            database=POSTGRES_DB,
            host=POSTGRES_HOST,
            port=POSTGRES_PORT,
            min_size=1,
            max_size=10,
        )
    return db_pool


logging.basicConfig(level=logging.INFO)

bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
dp = Dispatcher()

ADMIN_IDS_RAW = os.getenv("ADMIN_IDS", "620159705,1331701095,5273066461,1049352750")
ADMIN_USERS = {int(x.strip()) for x in ADMIN_IDS_RAW.split(",") if x.strip().isdigit()}
EXCLUDED_IDS = {14, 17, 22}

MSK = pytz.timezone("Europe/Moscow")

STAFF = [
    {"id": 100, "name": "Виктория Александровна", "tg_id": 1331701095, "vk_id": 233661166}
]

STUDENTS = [
    {"id": 1, "name": "Голубева Ольга", "tg_id": 7610841443, "vk_id": 1046066591},
    {"id": 2, "name": "Жуков Ярослав", "tg_id": 1693356589, "vk_id": 625235948},
    {"id": 3, "name": "Захарян Ангелина", "tg_id": 984245205, "vk_id": 503799617},
    {"id": 4, "name": "Исаев Исамутдин", "tg_id": 7312085971, "vk_id": 493372683},
    {"id": 5, "name": "Калашникова Виктория", "tg_id": 1145647467, "vk_id": 676760867},
    {"id": 6, "name": "Крюкова Екатерина", "tg_id": 5209067734, "vk_id": 653464614},
    {"id": 7, "name": "Лавренов Денис", "tg_id": 1776233614, "vk_id": 0},
    {"id": 8, "name": "Лапкин Никита", "tg_id": 786412327, "vk_id": 648972537},
    {"id": 9, "name": "Леваев Денис", "tg_id": 1380783132, "vk_id": 0},
    {"id": 10, "name": "Малюта Кирилл", "tg_id": 2036039791, "vk_id": 514896986},
    {"id": 11, "name": "Манин Даниил", "tg_id": 1426586903, "vk_id": 645040795},
    {"id": 12, "name": "Нестеренко Артем", "tg_id": 1590263622, "vk_id": 650095472},
    {"id": 13, "name": "Нестеренко Кирилл", "tg_id": 1816834428, "vk_id": 0},
    {"id": 14, "name": "Петровский Кирилл", "tg_id": 1049352750, "vk_id": 585815790},
    {"id": 15, "name": "Половинкин Максим", "tg_id": 5012979967, "vk_id": 651715517},
    {"id": 16, "name": "Попов Илья", "tg_id": 1678240030, "vk_id": 631499535},
    {"id": 17, "name": "Постнов Максим", "tg_id": 620159705, "vk_id": 437805680},
    {"id": 18, "name": "Резников Филипп", "tg_id": 1249491991, "vk_id": 607391313},
    {"id": 19, "name": "Скорик Глеб", "tg_id": 654109019, "vk_id": 509518532},
    {"id": 20, "name": "Филимонов Дмитрий", "tg_id": 6969927775, "vk_id": 604540036},
    {"id": 21, "name": "Франк Никита", "tg_id": 1329870096, "vk_id": 591025661},
    {"id": 22, "name": "Четвериков Вадим", "tg_id": 5273066461, "vk_id": 550484299},
]

ALL_PEOPLE = STUDENTS + STAFF
TG_NAME_MAP = {s["tg_id"]: s["name"] for s in ALL_PEOPLE if s.get("tg_id")}
VK_NAME_MAP = {s["vk_id"]: s["name"] for s in ALL_PEOPLE if s.get("vk_id") and s["vk_id"] != 0}

try:
    from app.data.schedule_data import BASE_SCHEDULE
except Exception:
    BASE_SCHEDULE = [
        # 1 семестр
        {"day": 0, "time": "09:40", "name": "Операционные системы и среды", "teacher": "Орлянская В.С.", "start": "2025-09-01", "end": "2025-10-25"},
        {"day": 0, "time": "11:20", "name": "Физическая культура", "teacher": "Иванова М.А.", "start": "2025-09-01", "end": "2025-10-25"},
        {"day": 0, "time": "13:10", "name": "Стандартизация, сертификация и тех. документоведение", "teacher": "Иванов Е.О.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 0, "time": "16:00", "name": "Основы проектирования баз данных", "teacher": "Петрушка С.А.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 1, "time": "08:00", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2025-09-01", "end": "2025-10-25"},
        {"day": 1, "time": "08:00", "name": "Основы философии", "teacher": "Еременко Я.А.", "start": "2025-10-27", "end": "2025-12-20"},
        {"day": 1, "time": "09:40", "name": "Элементы высшей математики", "teacher": "Сиканова В.А.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 1, "time": "11:20", "name": "Основы алгоритмизации и программирования", "teacher": "Сиканова В.А.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 1, "time": "13:10", "name": "Операционные системы и среды", "teacher": "Орлянская В.С.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 2, "time": "08:00", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2025-09-01", "end": "2025-10-25"},
        {"day": 2, "time": "08:00", "name": "Дискретная математика с элементами мат. логики", "teacher": "Корманенко Н.В.", "start": "2025-10-27", "end": "2025-12-20"},
        {"day": 2, "time": "09:40", "name": "Основы философии", "teacher": "Еременко Я.А.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 2, "time": "11:20", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 3, "time": "11:20", "name": "Информационные технологии", "teacher": "Кириченко Е.Г.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 3, "time": "13:10", "name": "Русский язык и культура речи", "teacher": "Безуленко Д.А.", "start": "2025-09-01", "end": "2025-10-25"},
        {"day": 3, "time": "13:10", "name": "Физическая культура", "teacher": "Иванова М.А.", "start": "2025-10-27", "end": "2025-12-20"},
        {"day": 3, "time": "14:50", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2025-09-01", "end": "2025-10-25"},
        {"day": 3, "time": "14:50", "name": "Операционные системы и среды", "teacher": "Орлянская В.С.", "start": "2025-10-27", "end": "2025-12-20"},
        {"day": 4, "time": "08:00", "name": "Стандартизация, сертификация и тех. документоведение", "teacher": "Иванов Е.О.", "start": "2025-01-09", "end": "2025-10-25"},
        {"day": 4, "time": "08:00", "name": "Основы проектирования баз данных", "teacher": "Петрушка С.А.", "start": "2025-10-27", "end": "2025-12-20"},
        {"day": 4, "time": "09:40", "name": "Архитектура аппаратных средств", "teacher": "Петрушка С.А.", "start": "2025-09-01", "end": "2025-12-20"},
        {"day": 4, "time": "11:20", "name": "Иностранный язык в проф. деятельности", "teacher": "Лабинцева Н.Г.", "start": "2025-01-09", "end": "2025-12-20"},
        {"day": 4, "time": "13:10", "name": "Дискретная математика с элементами мат. логики", "teacher": "Корманенко Н.В.", "start": "2025-01-09", "end": "2025-12-20"},
        # 2 семестр
        {"day": 0, "time": "09:10", "name": "Физическая культура", "teacher": "Иванова М.А.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 0, "time": "10:50", "name": "Элементы высшей математики", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 0, "time": "12:40", "name": "Основы алгоритмизации и программирования", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 0, "time": "14:20", "name": "Теория вероятностей и математическая статистика", "teacher": "Нечитайло М.С.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 1, "time": "08:00", "name": "Элементы высшей математики", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 1, "time": "09:40", "name": "Основы алгоритмизации и программирования", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 1, "time": "11:20", "name": "Основы алгоритмизации и программирования", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-02-14"},
        {"day": 1, "time": "11:20", "name": "История", "teacher": "Баркова А.С.", "start": "2026-02-16", "end": "2026-04-15"},
        {"day": 1, "time": "13:10", "name": "Иностранный язык в проф. деятельности", "teacher": "Лабинцева Н.Г.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 2, "time": "13:10", "name": "История", "teacher": "Баркова А.С.", "start": "2026-01-12", "end": "2026-02-28"},
        {"day": 2, "time": "13:10", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2026-03-02", "end": "2026-04-15"},
        {"day": 2, "time": "14:50", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 2, "time": "16:40", "name": "Экологические основы природопользования", "teacher": "Гадаева Д.М.", "start": "2026-02-16", "end": "2026-04-04"},
        {"day": 3, "time": "11:20", "name": "Основы алгоритмизации и программирования", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-01-24"},
        {"day": 3, "time": "11:20", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2026-01-26", "end": "2026-02-07"},
        {"day": 3, "time": "13:10", "name": "История", "teacher": "Баркова А.С.", "start": "2026-01-12", "end": "2026-02-14"},
        {"day": 3, "time": "13:10", "name": "Информационные системы", "teacher": "Кириченко Е.Е.", "start": "2026-03-09", "end": "2026-03-28"},
        {"day": 3, "time": "14:50", "name": "Информационные системы", "teacher": "Кириченко Е.Е.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 3, "time": "16:40", "name": "Информационные системы", "teacher": "Кириченко Е.Е.", "start": "2026-02-16", "end": "2026-03-07"},
        {"day": 4, "time": "09:40", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 4, "time": "11:20", "name": "МДК 11.01 Технология разработки и защиты баз данных", "teacher": "Сиканова В.А.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 4, "time": "13:10", "name": "Информационные системы", "teacher": "Кириченко Е.Е.", "start": "2026-01-12", "end": "2026-04-15"},
        {"day": 4, "time": "14:50", "name": "Информационные системы", "teacher": "Кириченко Е.Е.", "start": "2026-01-12", "end": "2026-03-21"},
    ]

UNDO_STORAGE = {}


async def notify_backend_duties():
    """Отправляет сигнал бэкенду, чтобы тот уведомил веб-интерфейс через WebSocket."""
    try:
        async with aiohttp.ClientSession() as session:
            await session.post(f"{BACKEND_URL}/internal/broadcast_duties", timeout=3)
    except Exception as e:
        logging.warning(f"Не удалось отправить broadcast на бэкенд: {e}")


# --- ГЕНЕРАЦИЯ EXCEL ---
async def generate_excel_report(year: str, month: str):
    y = int(year)
    m = int(month)
    month_prefix = f"{year}-{m:02d}-"
    _, last_day = calendar.monthrange(y, m)

    pool = await get_db_pool()

    # Загружаем посещаемость
    att_rows = await pool.fetch(
        "SELECT date, student_id, status FROM attendance WHERE date LIKE $1 AND status > 0",
        month_prefix + "%",
    )
    data = {}
    for row in att_rows:
        d_int = int(row["date"].split("-")[2])
        s_id = row["student_id"]
        if s_id not in data:
            data[s_id] = {}
        if d_int not in data[s_id]:
            data[s_id][d_int] = []
        data[s_id][d_int].append(row["status"])

    # Загружаем изменения в расписании
    ovr_rows = await pool.fetch(
        "SELECT date, time, is_canceled FROM overrides WHERE date LIKE $1",
        month_prefix + "%",
    )
    overrides_data = [(r["date"], r["time"], r["is_canceled"]) for r in ovr_rows]

    grey_days = set()
    for d in range(1, last_day + 1):
        date_str = f"{y}-{m:02d}-{d:02d}"
        wday = datetime(y, m, d).weekday()

        base_times = {
            l["time"]
            for l in BASE_SCHEDULE
            if l["day"] == wday and l["start"] <= date_str <= l["end"]
        }
        day_ovr = {r[1]: bool(r[2]) for r in overrides_data if r[0] == date_str}

        active_count = 0
        for t in base_times:
            if not day_ovr.get(t, False):
                active_count += 1
        for t, is_canc in day_ovr.items():
            if t not in base_times and not is_canc:
                active_count += 1

        if active_count == 0:
            grey_days.add(d)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    grey_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ru_months = [
        "",
        "января",
        "февраля",
        "марта",
        "апреля",
        "мая",
        "июня",
        "июля",
        "августа",
        "сентября",
        "октября",
        "ноября",
        "декабря",
    ]

    ws.column_dimensions["A"].width = 3
    START_COL = 2
    START_ROW = 3

    ws.cell(
        row=2,
        column=START_COL,
        value=f"ВЕДОМОСТЬ учета учебных часов за {ru_months[m]} {year} года",
    ).font = Font(size=14, bold=True)
    ws.merge_cells(
        start_row=2, start_column=START_COL, end_row=2, end_column=START_COL + last_day + 4
    )

    ws.cell(row=START_ROW, column=START_COL, value="№ п/п").font = bold_font
    ws.merge_cells(
        start_row=START_ROW,
        start_column=START_COL,
        end_row=START_ROW + 1,
        end_column=START_COL,
    )
    ws.column_dimensions[get_column_letter(START_COL)].width = 4
    ws.cell(row=START_ROW, column=START_COL + 1, value="ФИО").font = bold_font
    ws.merge_cells(
        start_row=START_ROW,
        start_column=START_COL + 1,
        end_row=START_ROW + 1,
        end_column=START_COL + 1,
    )
    ws.column_dimensions[get_column_letter(START_COL + 1)].width = 25

    for d in range(1, last_day + 1):
        col_idx = START_COL + 1 + d
        c = ws.cell(row=START_ROW, column=col_idx, value=d)
        c.font = bold_font
        ws.merge_cells(
            start_row=START_ROW,
            start_column=col_idx,
            end_row=START_ROW + 1,
            end_column=col_idx,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = 4
        if d in grey_days:
            c.fill = grey_fill

    tot_col = START_COL + last_day + 2
    ws.cell(row=START_ROW, column=tot_col, value="Итого").font = bold_font
    ws.merge_cells(
        start_row=START_ROW,
        start_column=tot_col,
        end_row=START_ROW + 1,
        end_column=tot_col,
    )
    ws.cell(row=START_ROW, column=tot_col + 1, value="Из них").font = bold_font
    ws.merge_cells(
        start_row=START_ROW,
        start_column=tot_col + 1,
        end_row=START_ROW,
        end_column=tot_col + 2,
    )
    ws.cell(row=START_ROW + 1, column=tot_col + 1, value="уваж.")
    ws.cell(row=START_ROW + 1, column=tot_col + 2, value="не уваж.")

    for r in range(START_ROW, START_ROW + 2):
        for col in range(START_COL, tot_col + 3):
            ws.cell(row=r, column=col).border = border

    row_idx = START_ROW + 2
    day_totals = {d: 0 for d in range(1, last_day + 1)}
    grand_total_nb = 0
    grand_total_uv = 0
    sorted_students = sorted(STUDENTS, key=lambda x: x["name"])

    for idx, s in enumerate(sorted_students, 1):
        ws.cell(row=row_idx, column=START_COL, value=idx).border = border
        ws.cell(row=row_idx, column=START_COL + 1, value=s["name"]).border = border
        s_nb = 0
        s_uv = 0
        for d in range(1, last_day + 1):
            cell = ws.cell(row=row_idx, column=START_COL + 1 + d)
            cell.border = border
            cell.alignment = center_align
            if s["id"] in data and d in data[s["id"]]:
                statuses = data[s["id"]][d]
                hours = len(statuses) * 2
                day_totals[d] += hours
                for st in statuses:
                    if st == 1:
                        s_nb += 2
                    if st == 2:
                        s_uv += 2
                cell.value = hours
                cell.fill = red_fill if 1 in statuses else yellow_fill
            elif d in grey_days:
                cell.fill = grey_fill

        grand_total_nb += s_nb
        grand_total_uv += s_uv
        ws.cell(row=row_idx, column=tot_col, value=s_nb + s_uv).border = border
        ws.cell(row=row_idx, column=tot_col + 1, value=s_uv).border = border
        ws.cell(row=row_idx, column=tot_col + 2, value=s_nb).border = border
        row_idx += 1

    ws.cell(row=row_idx, column=START_COL, value="Итого").font = bold_font
    ws.merge_cells(
        start_row=row_idx,
        start_column=START_COL,
        end_row=row_idx,
        end_column=START_COL + 1,
    )
    ws.cell(row=row_idx, column=START_COL).border = border
    ws.cell(row=row_idx, column=START_COL + 1).border = border
    for d in range(1, last_day + 1):
        c = ws.cell(row=row_idx, column=START_COL + 1 + d, value=day_totals[d] or 0)
        c.font = bold_font
        c.border = border

    ws.cell(row=row_idx, column=tot_col, value=grand_total_nb + grand_total_uv).font = bold_font
    ws.cell(row=row_idx, column=tot_col).border = border
    ws.cell(row=row_idx, column=tot_col + 1, value=grand_total_uv).font = bold_font
    ws.cell(row=row_idx, column=tot_col + 1).border = border
    ws.cell(row=row_idx, column=tot_col + 2, value=grand_total_nb).font = bold_font
    ws.cell(row=row_idx, column=tot_col + 2).border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- ПРИВЯЗКА СООБЩЕНИЙ (MESSAGE BRIDGE) ---
async def save_msg_link(tg_id: int, vk_id: int):
    pool = await get_db_pool()
    now = time.time()
    await pool.execute(
        """
        INSERT INTO message_bridge (tg_msg_id, vk_msg_id, created_at)
        VALUES ($1, $2, $3)
        ON CONFLICT (tg_msg_id) DO UPDATE SET vk_msg_id = $2, created_at = $3
        """,
        tg_id,
        vk_id,
        now,
    )


async def get_vk_by_tg(tg_id: int) -> int | None:
    pool = await get_db_pool()
    return await pool.fetchval("SELECT vk_msg_id FROM message_bridge WHERE tg_msg_id = $1", tg_id)


async def get_tg_by_vk(vk_id: int) -> int | None:
    pool = await get_db_pool()
    return await pool.fetchval("SELECT tg_msg_id FROM message_bridge WHERE vk_msg_id = $1", vk_id)


class AsyncVKBridge:
    def __init__(self):
        self.session = None
        self.user_cache = {}

    async def get_session(self):
        if not self.session or self.session.closed:
            self.session = aiohttp.ClientSession()
        return self.session

    async def api_call(self, method, params=None):
        if params is None:
            params = {}
        params["access_token"] = VK_TOKEN
        params["v"] = VK_API_VERSION
        session = await self.get_session()
        async with session.post(f"https://api.vk.com/method/{method}", data=params) as resp:
            return await resp.json()

    async def get_user_name(self, user_id):
        if user_id in VK_NAME_MAP:
            return VK_NAME_MAP[user_id]
        if user_id in self.user_cache:
            return self.user_cache[user_id]
        if user_id < 0:
            return "Группа"
        res = await self.api_call("users.get", {"user_ids": user_id})
        if "response" in res and res["response"]:
            user = res["response"][0]
            name = f"{user['first_name']} {user['last_name']}"
            self.user_cache[user_id] = name
            return name
        return str(user_id)

    async def get_users_info(self, user_ids: list):
        if not user_ids:
            return []
        ids_str = ",".join(map(str, user_ids))
        res = await self.api_call("users.get", {"user_ids": ids_str})
        return res.get("response", [])

    async def send_message(self, text, peer_id=VK_CHAT_PEER_ID, attachment=""):
        params = {
            "peer_id": peer_id,
            "message": text,
            "random_id": int(time.time() * 1000),
            "attachment": attachment,
        }
        res = await self.api_call("messages.send", params)
        return res.get("response")

    async def edit_message(self, vk_msg_id, peer_id, new_text, attachment=""):
        params = {
            "peer_id": peer_id,
            "message": new_text,
            "message_id": vk_msg_id,
            "attachment": attachment,
        }
        await self.api_call("messages.edit", params)

    async def get_attachments_str(self, vk_msg_id):
        res = await self.api_call("messages.getById", {"message_ids": vk_msg_id})
        items = res.get("response", {}).get("items", [])
        if not items:
            return ""
        atts = items[0].get("attachments", [])
        att_strs = []
        for a in atts:
            t = a["type"]
            if t in a:
                obj = a[t]
                att_str = f"{t}{obj.get('owner_id', '')}_{obj.get('id', '')}"
                if "access_key" in obj:
                    att_str += f"_{obj['access_key']}"
                att_strs.append(att_str)
        return ",".join(att_strs)

    async def upload_photo(self, photo_bytes, peer_id):
        try:
            server_res = await self.api_call("photos.getMessagesUploadServer", {"peer_id": peer_id})
            upload_url = server_res["response"]["upload_url"]

            session = await self.get_session()
            form = aiohttp.FormData()
            form.add_field("photo", photo_bytes, filename="photo.jpg", content_type="image/jpeg")
            async with session.post(upload_url, data=form) as resp:
                upload_data = await resp.json()

            save_res = await self.api_call(
                "photos.saveMessagesPhoto",
                {
                    "photo": upload_data["photo"],
                    "server": upload_data["server"],
                    "hash": upload_data["hash"],
                },
            )
            photo = save_res["response"][0]
            return f"photo{photo['owner_id']}_{photo['id']}"
        except Exception as e:
            logging.error(f"VK Photo Upload Error: {e}")
            return ""

    async def upload_document(self, file_bytes, filename, peer_id):
        try:
            server_res = await self.api_call(
                "docs.getMessagesUploadServer", {"type": "doc", "peer_id": peer_id}
            )
            upload_url = server_res["response"]["upload_url"]

            session = await self.get_session()
            form = aiohttp.FormData()
            form.add_field(
                "file", file_bytes, filename=filename, content_type="application/octet-stream"
            )

            async with session.post(upload_url, data=form) as resp:
                try:
                    upload_data = await resp.json()
                except Exception:
                    logging.error(f"ВК отклонил файл {filename}")
                    return "VK_REJECTED"

            save_res = await self.api_call(
                "docs.save", {"file": upload_data["file"], "title": filename}
            )

            if "response" in save_res:
                doc = save_res["response"]["doc"]
                return f"doc{doc['owner_id']}_{doc['id']}"
            return ""
        except Exception as e:
            logging.error(f"VK Doc Upload Error: {e}")
            return ""

    async def upload_voice(self, voice_bytes, peer_id):
        try:
            server_res = await self.api_call(
                "docs.getMessagesUploadServer", {"type": "audio_message", "peer_id": peer_id}
            )
            upload_url = server_res["response"]["upload_url"]

            session = await self.get_session()
            form = aiohttp.FormData()
            form.add_field("file", voice_bytes, filename="voice.ogg", content_type="audio/ogg")
            async with session.post(upload_url, data=form) as resp:
                upload_data = await resp.json()

            save_res = await self.api_call("docs.save", {"file": upload_data["file"]})
            audio_msg = save_res["response"]["audio_message"]
            return f"audio_message{audio_msg['owner_id']}_{audio_msg['id']}"
        except Exception as e:
            logging.error(f"VK Voice Upload Error: {e}")
            return ""

    async def download_file_to_bytes(self, url: str):
        try:
            session = await self.get_session()
            async with session.get(url) as resp:
                if resp.status == 200:
                    return await resp.read()
        except Exception as e:
            logging.error(f"Ошибка скачивания файла: {e}")
        return None

    async def format_fwd_messages(self, fwd_list, level=1):
        res = []
        prefix = "┃ " * level
        for fwd in fwd_list:
            author = await self.get_user_name(fwd.get("from_id", 0))
            text = fwd.get("text", "")

            atts = fwd.get("attachments", [])
            att_labels = []
            if atts:
                for a in atts:
                    if a["type"] == "photo":
                        att_labels.append("🖼 Фото")
                    elif a["type"] == "doc":
                        att_labels.append("📄 Файл")
                    elif a["type"] == "audio_message":
                        att_labels.append("🎤 Голос")
                    elif a["type"] == "video":
                        att_labels.append("🎥 Видео")

            att_str = f"<i>({', '.join(att_labels)})</i>" if att_labels else ""
            content = f"{html.escape(text)} {att_str}".strip()
            if not content:
                content = "<i>[Пустое сообщение]</i>"

            res.append(f"{prefix}<b>{html.escape(author)}</b>: {content}")

            if "fwd_messages" in fwd and fwd["fwd_messages"]:
                inner_fwd = await self.format_fwd_messages(fwd["fwd_messages"], level + 1)
                res.extend(inner_fwd)
        return res

    async def collect_all_attachments(self, msg):
        all_atts = []
        if "attachments" in msg:
            all_atts.extend(msg["attachments"])
        if "fwd_messages" in msg:
            for fwd in msg["fwd_messages"]:
                all_atts.extend(await self.collect_all_attachments(fwd))
        return all_atts

    async def start_longpoll(self):
        if not VK_TOKEN:
            return
        logging.info("VK LongPoll запущен")

        while True:
            try:
                server_data = await self.api_call(
                    "groups.getLongPollServer", {"group_id": VK_GROUP_ID}
                )
                server = server_data["response"]["server"]
                key = server_data["response"]["key"]
                ts = server_data["response"]["ts"]

                session = await self.get_session()

                while True:
                    async with session.get(
                        f"{server}?act=a_check&key={key}&ts={ts}&wait=25", timeout=30
                    ) as resp:
                        r = await resp.json()

                    if "failed" in r:
                        break
                    ts = r["ts"]

                    for update in r.get("updates", []):
                        if update["type"] == "message_new":
                            msg = update["object"]["message"]
                            if msg.get("from_id") == -VK_GROUP_ID:
                                continue

                            vk_peer_id = msg.get("peer_id")
                            tg_target_chat = CHAT_MAP.get(vk_peer_id)
                            if not tg_target_chat:
                                continue

                            vk_msg_hash = msg.get("conversation_message_id", msg.get("id"))
                            if vk_msg_hash in PROCESSED_VK_MSGS:
                                continue
                            PROCESSED_VK_MSGS.add(vk_msg_hash)

                            if len(PROCESSED_VK_MSGS) > 1000:
                                PROCESSED_VK_MSGS.clear()

                            async def process_delayed_vk_msg(original_msg, target_chat, delay):
                                await asyncio.sleep(delay)
                                current_msg = original_msg
                                try:
                                    fresh_res = None
                                    vk_id = original_msg.get("id", 0)
                                    conv_id = original_msg.get("conversation_message_id")

                                    if vk_id != 0:
                                        fresh_res = await self.api_call(
                                            "messages.getById", {"message_ids": vk_id}
                                        )
                                    elif conv_id:
                                        fresh_res = await self.api_call(
                                            "messages.getByConversationMessageId",
                                            {
                                                "peer_id": original_msg.get("peer_id"),
                                                "conversation_message_ids": conv_id,
                                            },
                                        )

                                    if (
                                        fresh_res
                                        and "response" in fresh_res
                                        and fresh_res["response"].get("items")
                                    ):
                                        current_msg = fresh_res["response"]["items"][0]
                                except Exception as e:
                                    logging.error(f"Ошибка при обновлении сообщения ВК: {e}")

                                text = current_msg.get("text", "")
                                name = await self.get_user_name(current_msg.get("from_id", 0))

                                user_url = (
                                    f"https://vk.com/id{current_msg.get('from_id')}"
                                    if current_msg.get("from_id", 0) > 0
                                    else "https://vk.com/club"
                                    + str(abs(current_msg.get("from_id", 0)))
                                )
                                tg_text_parts = [
                                    f"🔵 <b>[VK]</b> <a href='{user_url}'>{html.escape(name)}</a>"
                                ]

                                if "reply_message" in current_msg:
                                    reply_msg = current_msg["reply_message"]
                                    reply_author = await self.get_user_name(
                                        reply_msg.get("from_id", 0)
                                    )
                                    orig_text = reply_msg.get("text", "")
                                    if not orig_text and reply_msg.get("attachments"):
                                        orig_text = "[Вложение]"

                                    tg_text_parts.append(
                                        f"<blockquote expandable><b>👤 В ответ {html.escape(reply_author)}:</b>\n<i>{html.escape(orig_text)}</i></blockquote>"
                                    )

                                if "fwd_messages" in current_msg and current_msg["fwd_messages"]:
                                    fwd_lines = await self.format_fwd_messages(
                                        current_msg["fwd_messages"]
                                    )
                                    tg_text_parts.append(
                                        "<b>✉️ Пересланные сообщения:</b>\n"
                                        + "\n".join(fwd_lines)
                                    )

                                if text:
                                    tg_text_parts.append(html.escape(text))

                                tg_text = "\n\n".join(tg_text_parts)

                                all_attachments = await self.collect_all_attachments(current_msg)
                                photos = []
                                docs = []
                                voices = []

                                for att in all_attachments:
                                    try:
                                        if att["type"] == "photo":
                                            sizes = att["photo"]["sizes"]
                                            photo_url = sorted(sizes, key=lambda x: x["width"])[
                                                -1
                                            ]["url"]
                                            file_bytes = await self.download_file_to_bytes(
                                                photo_url
                                            )
                                            if file_bytes:
                                                photos.append(
                                                    BufferedInputFile(
                                                        file_bytes,
                                                        filename=f"photo_{uuid.uuid4().hex[:6]}.jpg",
                                                    )
                                                )

                                        elif att["type"] == "doc":
                                            doc_url = att["doc"]["url"]
                                            doc_title = att["doc"].get(
                                                "title", f"file_{uuid.uuid4().hex[:6]}"
                                            )
                                            file_bytes = await self.download_file_to_bytes(doc_url)
                                            if file_bytes:
                                                docs.append(
                                                    BufferedInputFile(
                                                        file_bytes, filename=doc_title
                                                    )
                                                )

                                        elif att["type"] == "audio_message":
                                            voice_url = att["audio_message"]["link_ogg"]
                                            file_bytes = await self.download_file_to_bytes(
                                                voice_url
                                            )
                                            if file_bytes:
                                                voices.append(
                                                    BufferedInputFile(
                                                        file_bytes,
                                                        filename=f"voice_{uuid.uuid4().hex[:6]}.ogg",
                                                    )
                                                )

                                        elif att["type"] == "video":
                                            v = att["video"]
                                            video_url = (
                                                f"https://vk.com/video{v['owner_id']}_{v['id']}"
                                            )
                                            if "access_key" in v:
                                                video_url += f"_{v['access_key']}"
                                            tg_text += f"\n\n🎥 <a href='{video_url}'><b>Видеозапись ВКонтакте</b></a>"
                                    except Exception as e:
                                        logging.error(f"Ошибка обработки вложения: {e}")

                                try:
                                    sent_msgs = []
                                    caption_used = False

                                    if photos:
                                        for i in range(0, len(photos), 10):
                                            chunk = photos[i : i + 10]
                                            if len(chunk) == 1:
                                                m = await bot.send_photo(
                                                    target_chat,
                                                    photo=chunk[0],
                                                    caption=tg_text if not caption_used else None,
                                                )
                                                sent_msgs.append(m)
                                            else:
                                                media = [
                                                    InputMediaPhoto(
                                                        media=p,
                                                        caption=tg_text
                                                        if idx == 0 and not caption_used
                                                        else None,
                                                    )
                                                    for idx, p in enumerate(chunk)
                                                ]
                                                m_list = await bot.send_media_group(
                                                    target_chat, media=media
                                                )
                                                sent_msgs.extend(m_list)
                                            caption_used = True

                                    if docs:
                                        for d in docs:
                                            m = await bot.send_document(
                                                target_chat,
                                                document=d,
                                                caption=tg_text if not caption_used else None,
                                            )
                                            sent_msgs.append(m)
                                            caption_used = True

                                    if voices:
                                        for v in voices:
                                            m = await bot.send_voice(
                                                target_chat,
                                                voice=v,
                                                caption=tg_text if not caption_used else None,
                                            )
                                            sent_msgs.append(m)
                                            caption_used = True

                                    if not caption_used:
                                        m = await bot.send_message(target_chat, tg_text)
                                        sent_msgs.append(m)

                                    if sent_msgs:
                                        vk_save_id = current_msg.get("id")
                                        if vk_save_id == 0 or vk_save_id is None:
                                            vk_save_id = current_msg.get("conversation_message_id")

                                        if vk_save_id:
                                            await save_msg_link(sent_msgs[0].message_id, vk_save_id)

                                except Exception as e:
                                    logging.error(f"Error sending to TG: {e}")

                            has_attachments = any(
                                [
                                    msg.get("attachments"),
                                    msg.get("fwd_messages"),
                                    msg.get("reply_message"),
                                ]
                            )

                            wait_time = 4.5 if has_attachments else 0.1
                            asyncio.create_task(
                                process_delayed_vk_msg(msg, tg_target_chat, wait_time)
                            )

                        elif update["type"] == "message_edit":
                            msg_obj = update["object"]
                            msg = msg_obj.get("message", msg_obj)

                            if msg.get("from_id") == -VK_GROUP_ID:
                                continue

                            vk_peer_id = msg.get("peer_id")
                            tg_target_chat = CHAT_MAP.get(vk_peer_id)
                            if not tg_target_chat:
                                continue

                            vk_search_id = msg.get("id")
                            if vk_search_id == 0 or vk_search_id is None:
                                vk_search_id = msg.get("conversation_message_id")

                            if not vk_search_id:
                                continue

                            tg_msg_id = await get_tg_by_vk(vk_search_id)
                            if tg_msg_id:
                                name = await self.get_user_name(msg.get("from_id", 0))
                                text = msg.get("text", "")

                                tg_text_parts = [
                                    f"🔵 <b>[VK] {html.escape(name)}</b> <i>(изменено)</i>"
                                ]

                                if "reply_message" in msg:
                                    reply_msg = msg["reply_message"]
                                    reply_author = await self.get_user_name(
                                        reply_msg.get("from_id", 0)
                                    )
                                    orig_text = reply_msg.get("text", "")
                                    if not orig_text and reply_msg.get("attachments"):
                                        orig_text = "[Вложение]"

                                    tg_text_parts.append(
                                        f"<blockquote expandable><b>👤 В ответ {html.escape(reply_author)}:</b>\n<i>{html.escape(orig_text)}</i></blockquote>"
                                    )

                                if "fwd_messages" in msg and msg["fwd_messages"]:
                                    fwd_lines = await self.format_fwd_messages(msg["fwd_messages"])
                                    tg_text_parts.append(
                                        "<b>✉️ Пересланные сообщения:</b>\n"
                                        + "\n".join(fwd_lines)
                                    )

                                if text:
                                    tg_text_parts.append(html.escape(text))

                                tg_text = "\n\n".join(tg_text_parts)

                                all_attachments = await self.collect_all_attachments(msg)
                                for att in all_attachments:
                                    if att["type"] == "video":
                                        v = att["video"]
                                        video_url = (
                                            f"https://vk.com/video{v['owner_id']}_{v['id']}"
                                        )
                                        if "access_key" in v:
                                            video_url += f"_{v['access_key']}"
                                        tg_text += f"\n\n🎥 <a href='{video_url}'><b>Видеозапись ВКонтакте</b></a>"

                                try:
                                    await bot.edit_message_caption(
                                        chat_id=tg_target_chat,
                                        message_id=tg_msg_id,
                                        caption=tg_text,
                                    )
                                except Exception as e_cap:
                                    try:
                                        await bot.edit_message_text(
                                            tg_text,
                                            chat_id=tg_target_chat,
                                            message_id=tg_msg_id,
                                        )
                                    except Exception as e_text:
                                        logging.error(
                                            f"Не удалось обновить сообщение в ТГ. Ошибка подписи: {e_cap} | Ошибка текста: {e_text}"
                                        )
                            else:
                                logging.warning(
                                    f"Редактирование ВК: не найдена привязка ТГ для сообщения ВК ID {msg.get('id')}"
                                )

            except Exception as e:
                logging.error(f"VK LongPoll Error: {e}")
                await asyncio.sleep(5)


vk_bridge = AsyncVKBridge()


@dp.message(Command("clear_media"))
async def handle_clear_media(message: Message):
    if message.from_user.id not in ADMIN_USERS:
        return
    if not message.reply_to_message:
        return await message.answer("⚠️ Ответь на сообщение, с которого начать.")

    start_id = message.reply_to_message.message_id
    chat_id = message.chat.id
    vk_peer_id = REVERSE_CHAT_MAP.get(chat_id)
    if not vk_peer_id:
        return await message.answer("❌ Чат не сопоставлен с VK.")

    pool = await get_db_pool()
    rows = await pool.fetch(
        "SELECT tg_msg_id, vk_msg_id FROM message_bridge WHERE tg_msg_id >= $1 ORDER BY tg_msg_id ASC",
        start_id,
    )

    if not rows:
        return await message.answer("❌ Сообщения не найдены в базе.")
    status = await message.answer(f"⏳ Проверяю {len(rows)} сообщений...")

    success_count = 0
    for r in rows:
        tg_mid = r["tg_msg_id"]
        vk_id = r["vk_msg_id"]
        if not vk_id:
            continue
        try:
            vk_res = await vk_bridge.api_call(
                "messages.getByConversationMessageId",
                {"peer_id": vk_peer_id, "conversation_message_ids": vk_id},
            )
            items = vk_res.get("response", {}).get("items", [])
            if not items:
                vk_res = await vk_bridge.api_call(
                    "messages.getById", {"message_ids": vk_id}
                )
                items = vk_res.get("response", {}).get("items", [])

            if not items:
                continue
            v = items[0]
            f_id = v.get("from_id", 0)

            if f_id < 0:
                continue

            name = await vk_bridge.get_user_name(f_id)
            user_url = f"https://vk.com/id{f_id}"

            parts = [f"🔵 <b>[VK]</b> <a href='{user_url}'>{html.escape(name)}</a>"]

            if "reply_message" in v:
                r = v["reply_message"]
                rn = await vk_bridge.get_user_name(r.get("from_id", 0))
                rt = r.get("text") or "[Вложение]"
                parts.append(
                    f"<blockquote expandable><b>👤 В ответ {html.escape(rn)}:</b>\n<i>{html.escape(rt)}</i></blockquote>"
                )

            if v.get("text"):
                parts.append(html.escape(v["text"]))

            final_text = "\n\n".join(parts)

            await bot.send_message(chat_id=chat_id, text=final_text, disable_notification=True)
            await bot.delete_message(chat_id, tg_mid)
            success_count += 1

            await asyncio.sleep(0.3)
        except Exception as e:
            logging.error(f"Ошибка при замене {tg_mid}: {e}")

    await status.delete()
    sent_ok = await message.answer(f"✅ Очищено сообщений из ВК: {success_count}.")
    asyncio.create_task(delete_later([message, sent_ok], 5))


def parse_date_input(date_str: str) -> str | None:
    date_str = date_str.strip()
    current_year = datetime.now(MSK).year
    patterns = [
        ("%Y-%m-%d", False),
        ("%d.%m.%Y", False),
        ("%d.%m.%y", False),
        ("%d.%m", True),
        ("%d/%m/%Y", False),
    ]
    for fmt, add_year in patterns:
        try:
            dt = datetime.strptime(date_str, fmt)
            if add_year:
                dt = dt.replace(year=current_year)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


async def delete_later(messages: list[Message], delay: int = 60):
    await asyncio.sleep(delay)
    for msg in messages:
        with suppress(Exception):
            await msg.delete()


@dp.message(CommandStart())
async def handle_start(message: Message):
    args_list = message.text.split()
    args = args_list[1] if len(args_list) > 1 else None

    if args and args.startswith("report_"):
        if message.from_user.id not in ADMIN_USERS:
            return await message.answer("⛔️ Только администраторы могут скачивать отчеты.")
        try:
            _, year, month = args.split("_")
            if not (2020 < int(year) < 2030 and 0 < int(month) < 13):
                raise ValueError()
            await message.answer("⏳ Формирую отчет... Пожалуйста, подождите.")
            await bot.send_chat_action(chat_id=message.chat.id, action="upload_document")
            excel_file = await generate_excel_report(year, month)
            filename = f"Ведомость_{year}_{month}.xlsx"
            document = BufferedInputFile(excel_file.getvalue(), filename=filename)
            await bot.send_document(
                chat_id=message.chat.id,
                document=document,
                caption=f"✅ Ваш отчет за {month}-{year} готов!",
            )
        except Exception as e:
            await message.answer(f"⚠️ Не удалось создать отчет. Ошибка: {e}")
        return

    await message.answer(
        "👋 <b>Привет!</b> Я бот для управления графиком дежурств.\nНажми /help чтобы узнать доступные команды."
    )


@dp.message(Command("help"))
async def help_command(message: Message):
    text = (
        "📖 <b>СПРАВОЧНАЯ ИНФОРМАЦИЯ</b>\n━━━━━━━━━━━━━━━━━━\n\n"
        "🟢 <b>ДОСТУПНО ВСЕМ:</b>\n"
        "🔹 <code>/list</code> — Посмотреть график дежурств.\n"
        "🟠 <b>ТОЛЬКО ДЛЯ АДМИНИСТРАТОРОВ:</b>\n"
        "🔸 <code>/dury [дата]</code> — Назначить дежурных.\n"
    )
    sent_msg = await message.answer(text, disable_notification=True)
    asyncio.create_task(delete_later([message, sent_msg], delay=60))


@dp.message(Command("list"))
async def handle_list(message: Message):
    pool = await get_db_pool()
    rows = await pool.fetch("SELECT student_id, date FROM duties")
    db_data = {r["student_id"]: r["date"] for r in rows}

    lines = message.text.split("\n")
    excluded_names = [l.strip().lower() for l in lines[1:] if l.strip()]

    duty_list = []
    for s in STUDENTS:
        if s["id"] in EXCLUDED_IDS:
            continue
        if any(ex in s["name"].lower() for ex in excluded_names):
            continue
        duty_list.append({"name": s["name"], "date": db_data.get(s["id"])})

    duty_list.sort(key=lambda x: (x["date"] is None, x["date"]))

    target_count = 3
    current_count = 0
    red_date = None
    blue_dates = set()
    dated_people = [x for x in duty_list if x["date"]]

    if dated_people:
        red_date = dated_people[0]["date"]
        for p in dated_people:
            d = p["date"]
            if d == red_date:
                current_count += 1
            else:
                if current_count < target_count:
                    blue_dates.add(d)
                    current_count += 1
                else:
                    if d in blue_dates:
                        current_count += 1
                    else:
                        break

    text = "📋 <b>ГРАФИК ДЕЖУРСТВ</b>\n━━━━━━━━━━━━━━━━━━\n"
    for entry in duty_list:
        mark = "⚪"
        d_val = entry["date"]
        if d_val:
            dt_str = (
                f"<code>{datetime.strptime(d_val, '%Y-%m-%d').strftime('%d.%m.%Y')}</code>"
            )
            if d_val == red_date:
                mark = "🔴"
            elif d_val in blue_dates:
                mark = "🔵"
        else:
            dt_str = "<code>   —      </code>"
        text += f"{mark} <b>{html.escape(entry['name'])}</b>\n      └ {dt_str}\n"

    sent = await message.answer(text, disable_notification=True)
    asyncio.create_task(delete_later([message, sent], 60))


@dp.message(Command("vk_names"))
async def handle_vk_names(message: Message):
    if message.from_user.id not in ADMIN_USERS:
        return await message.answer("⛔️ У вас нет прав.")

    sent_status = await message.answer("🔍 Синхронизация со списками VK...")

    with_vk_id = [s for s in ALL_PEOPLE if s.get("vk_id") and s["vk_id"] != 0]
    no_vk_id = [s for s in ALL_PEOPLE if not s.get("vk_id") or s["vk_id"] == 0]

    found_list = []
    error_id_list = []

    try:
        if with_vk_id:
            vk_ids = [s["vk_id"] for s in with_vk_id]
            vk_users = await vk_bridge.get_users_info(vk_ids)
            vk_response_map = {
                u["id"]: f"{u['first_name']} {u['last_name']}" for u in vk_users
            }

            for s in with_vk_id:
                v_id = s["vk_id"]
                if v_id in vk_response_map:
                    found_list.append(f"✅ {s['name']} → <b>{vk_response_map[v_id]}</b>")
                else:
                    error_id_list.append(f"❌ {s['name']} (ID: <code>{v_id}</code>)")

        report = ["📊 <b>ПРОВЕРКА СПИСКА VK</b>\n"]

        if found_list:
            report.append("<b>Успешно сопоставлены:</b>")
            report.extend(found_list)
            report.append("")

        if error_id_list:
            report.append("<b>Неверные ID (профиль не найден):</b>")
            report.extend(error_id_list)
            report.append("")

        if no_vk_id:
            report.append("<b>ID не указан в коде (vk_id: 0):</b>")
            for s in no_vk_id:
                report.append(f"⚠️ {s['name']}")

        await sent_status.edit_text("\n".join(report))

    except Exception as e:
        logging.error(f"Ошибка проверки имен: {e}")
        await sent_status.edit_text(f"❌ Ошибка API: {e}")


@dp.message(Command("dury"))
async def handle_dury(message: Message):
    if message.from_user.id not in ADMIN_USERS:
        return

    lines = message.text.split("\n")
    cmd_parts = lines[0].split()
    target_date = datetime.now(MSK).strftime("%Y-%m-%d")

    if len(cmd_parts) > 1:
        parsed = parse_date_input(cmd_parts[1])
        if parsed:
            target_date = parsed
        else:
            return await message.answer("⚠️ Неверный формат даты.")

    display_date = datetime.strptime(target_date, "%Y-%m-%d").strftime("%d.%m.%Y")
    names_input = lines[1:]

    updated_ids = []
    not_found = []
    undo_list = []

    pool = await get_db_pool()
    rows = await pool.fetch("SELECT student_id, date FROM duties")
    current_state = {r["student_id"]: r["date"] for r in rows}

    for line in names_input:
        search = line.strip().lower()
        if not search:
            continue
        found_student = next(
            (s for s in STUDENTS if search in s["name"].lower()), None
        )

        if found_student:
            s_id = found_student["id"]
            undo_list.append({"id": s_id, "old_date": current_state.get(s_id)})
            await pool.execute(
                """
                INSERT INTO duties (student_id, date) VALUES ($1, $2)
                ON CONFLICT (student_id) DO UPDATE SET date = $2
                """,
                s_id,
                target_date,
            )
            updated_ids.append(found_student["name"])
        else:
            not_found.append(line.strip())

    if not updated_ids:
        return await message.answer(
            f"🤷‍♂️ Никого не нашел.\nНе найдены: {', '.join(not_found)}"
        )

    # Уведомляем веб-клиентов
    asyncio.create_task(notify_backend_duties())

    undo_id = str(uuid.uuid4())[:8]
    UNDO_STORAGE[undo_id] = undo_list
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="↩️ Отменить", callback_data=f"undo_dury:{undo_id}")]
        ]
    )

    resp = (
        f"🔔 <b>Назначены дежурные!</b>\n📅 Дата: <code>{display_date}</code>\n━━━━━━━━━━━━━━━━━━\n"
    )
    for name in updated_ids:
        resp += f"✅ <b>{html.escape(name)}</b>\n"
    if not_found:
        resp += f"\n⚠️ <b>Не найдены:</b>\n" + "\n".join(
            [html.escape(n) for n in not_found]
        )
    resp += f"\n\n👤 <b>Назначил:</b> {message.from_user.mention_html()}"

    await message.answer(resp, reply_markup=kb)
    clean_vk_text = re.sub(r"<[^>]+>", "", resp)
    await vk_bridge.send_message(clean_vk_text)


@dp.callback_query(F.data.startswith("undo_dury:"))
async def process_undo(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_USERS:
        return await callback.answer("⛔️ Нет прав", show_alert=True)

    undo_id = callback.data.split(":")[1]
    undo_data = UNDO_STORAGE.get(undo_id)
    if not undo_data:
        return await callback.answer("⏳ Время истекло", show_alert=True)

    pool = await get_db_pool()
    for item in undo_data:
        if item["old_date"] is None:
            await pool.execute("DELETE FROM duties WHERE student_id = $1", item["id"])
        else:
            await pool.execute(
                """
                INSERT INTO duties (student_id, date) VALUES ($1, $2)
                ON CONFLICT (student_id) DO UPDATE SET date = $2
                """,
                item["id"],
                item["old_date"],
            )

    del UNDO_STORAGE[undo_id]
    await callback.message.edit_text(
        "🔄 <b>ИЗМЕНЕНИЯ ОТМЕНЕНЫ!</b>\nВозвращены старые даты.", reply_markup=None
    )
    await callback.answer("Готово")
    await vk_bridge.send_message("🔄 Назначение дежурных отменено.")
    asyncio.create_task(notify_backend_duties())


@dp.callback_query(F.data.startswith("web_undo:"))
async def process_web_undo(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_USERS:
        return await callback.answer("⛔️ Нет прав", show_alert=True)

    undo_id = callback.data.split(":")[1]
    pool = await get_db_pool()
    row = await pool.fetchrow(
        "SELECT data FROM web_undos WHERE undo_id = $1", undo_id
    )

    restored = False
    if row:
        undo_data = json.loads(row["data"])
        for item in undo_data:
            if item.get("date") is None:
                await pool.execute(
                    "DELETE FROM duties WHERE student_id = $1", item["id"]
                )
            else:
                await pool.execute(
                    """
                    INSERT INTO duties (student_id, date) VALUES ($1, $2)
                    ON CONFLICT (student_id) DO UPDATE SET date = $2
                    """,
                    item["id"],
                    item["date"],
                )
        await pool.execute("DELETE FROM web_undos WHERE undo_id = $1", undo_id)
        restored = True

    if restored:
        await callback.message.edit_text(
            f"{callback.message.html_text}\n\n❌ <b>ОТМЕНЕНО</b>", reply_markup=None
        )
        await callback.answer("Изменения отменены")
        await vk_bridge.send_message("🔄 Назначение дежурных с сайта отменено.")
        asyncio.create_task(notify_backend_duties())
    else:
        await callback.answer("⏳ Невозможно отменить", show_alert=True)
        await callback.message.edit_reply_markup(reply_markup=None)


# === БУФЕР АЛЬБОМОВ ДЛЯ ОТПРАВКИ ИЗ ТГ В ВК ===
ALBUM_CACHE = {}


async def process_tg_messages_to_vk(messages: list[Message], target_vk_peer: int):
    first_msg = messages[0]
    tg_user_id = first_msg.from_user.id
    name = TG_NAME_MAP.get(tg_user_id, first_msg.from_user.full_name)

    if first_msg.from_user.username:
        tg_link = first_msg.from_user.username
        vk_text_parts = [f"✈️ [TG] {name} ({tg_link})"]
    else:
        vk_text_parts = [f"✈️ [TG] {name}"]

    text_content = ""
    for m in messages:
        if m.text or m.caption:
            text_content = m.text or m.caption
            break

    if first_msg.reply_to_message:
        reply_user = (
            first_msg.reply_to_message.from_user.full_name
            if first_msg.reply_to_message.from_user
            else "Кого-то"
        )
        orig_text = (
            first_msg.reply_to_message.text
            or first_msg.reply_to_message.caption
            or "[Медиафайл/Вложение]"
        )

        if len(orig_text) > 100:
            display_text = orig_text[:97] + "..."
        else:
            display_text = orig_text

        quote_block = f"┃ 👤 {reply_user}\n┃ 💬 {display_text}"
        vk_text_parts.append(quote_block)

    attachments = []

    async def safe_download(file_id):
        try:
            file_info = await bot.get_file(file_id)
            return (await bot.download_file(file_info.file_path)).read()
        except Exception as ex:
            if "file is too big" in str(ex).lower():
                return "TOO_BIG"
            logging.warning(f"Ошибка скачивания ТГ: {ex}")
            return None

    def handle_vk_upload(att_result, fname, icon):
        if att_result == "VK_REJECTED":
            vk_text_parts.append(
                f"{icon} <i>[ВКонтакте отклонил файл {fname} (формат запрещен в ВК)]</i>"
            )
        elif att_result:
            attachments.append(att_result)

    for msg in messages:
        try:
            if msg.photo:
                photo_bytes = await safe_download(msg.photo[-1].file_id)
                if isinstance(photo_bytes, bytes):
                    att = await vk_bridge.upload_photo(photo_bytes, target_vk_peer)
                    if att:
                        attachments.append(att)

            elif msg.document:
                doc_bytes = await safe_download(msg.document.file_id)
                fname = msg.document.file_name or f"file_{uuid.uuid4().hex[:6]}"
                if isinstance(doc_bytes, bytes):
                    att = await vk_bridge.upload_document(
                        doc_bytes, fname, target_vk_peer
                    )
                    handle_vk_upload(att, fname, "📄")
                elif doc_bytes == "TOO_BIG":
                    vk_text_parts.append(
                        f"📄 <i>[Файл {fname} больше 20МБ (лимит Telegram бота)]</i>"
                    )

            elif msg.audio:
                audio_bytes = await safe_download(msg.audio.file_id)
                fname = msg.audio.file_name or "audio.mp3"
                if isinstance(audio_bytes, bytes):
                    att = await vk_bridge.upload_document(
                        audio_bytes, fname, target_vk_peer
                    )
                    handle_vk_upload(att, fname, "🎵")
                elif audio_bytes == "TOO_BIG":
                    vk_text_parts.append(f"🎵 <i>[Аудиофайл {fname} больше 20МБ]</i>")
                if not text_content:
                    text_content = "🎵 [Аудиофайл]"

            elif msg.voice:
                voice_bytes = await safe_download(msg.voice.file_id)
                if isinstance(voice_bytes, bytes):
                    att = await vk_bridge.upload_voice(voice_bytes, target_vk_peer)
                    if att:
                        attachments.append(att)
                if not text_content:
                    text_content = "🎤"

            elif msg.video:
                video_bytes = await safe_download(msg.video.file_id)
                fname = msg.video.file_name or f"video_{uuid.uuid4().hex[:6]}.mp4"
                if isinstance(video_bytes, bytes):
                    att = await vk_bridge.upload_document(
                        video_bytes, fname, target_vk_peer
                    )
                    handle_vk_upload(att, fname, "🎥")
                elif video_bytes == "TOO_BIG":
                    vk_text_parts.append(
                        f"🎥 <i>[Видео {fname} больше 20МБ (лимит Telegram бота)]</i>"
                    )

            elif msg.sticker:
                sticker_file = await bot.get_file(msg.sticker.file_id)
                sticker_bytes = (await bot.download_file(sticker_file.file_path)).read()
                fname = f"sticker_{uuid.uuid4().hex[:6]}.webp"
                att = await vk_bridge.upload_document(
                    sticker_bytes, fname, target_vk_peer
                )
                if att:
                    attachments.append(att)
                    emoji = msg.sticker.emoji or "✨"
                    if not text_content:
                        text_content = f"{emoji} [Стикер из Telegram]"

            elif msg.video_note:
                if not text_content:
                    text_content = "[Видеосообщение ⭕️]"

        except Exception as e:
            logging.error(f"Не удалось обработать медиа из ТГ: {e}")

    if text_content:
        vk_text_parts.append(text_content)

    vk_text = "\n\n".join(vk_text_parts)
    attachment_str = ",".join(attachments)

    vk_msg_id = await vk_bridge.send_message(
        vk_text, target_vk_peer, attachment_str
    )

    if vk_msg_id:
        for m in messages:
            await save_msg_link(m.message_id, vk_msg_id)


@dp.message(F.chat.id.in_(REVERSE_CHAT_MAP.keys()), ~F.text.startswith("/"))
async def tg_to_vk_handler(message: Message):
    if message.from_user.is_bot:
        return

    target_vk_peer = REVERSE_CHAT_MAP.get(message.chat.id)
    if not target_vk_peer:
        return

    if message.media_group_id:
        if message.media_group_id not in ALBUM_CACHE:
            ALBUM_CACHE[message.media_group_id] = []

            async def wait_and_process(peer_id):
                await asyncio.sleep(1.5)
                msgs = ALBUM_CACHE.pop(message.media_group_id, [])
                if msgs:
                    await process_tg_messages_to_vk(msgs, peer_id)

            asyncio.create_task(wait_and_process(target_vk_peer))

        ALBUM_CACHE[message.media_group_id].append(message)
    else:
        await process_tg_messages_to_vk([message], target_vk_peer)


@dp.edited_message(F.chat.id.in_(REVERSE_CHAT_MAP.keys()))
async def tg_edit_to_vk_handler(message: Message):
    if message.from_user.is_bot:
        return

    target_vk_peer = REVERSE_CHAT_MAP.get(message.chat.id)
    if not target_vk_peer:
        return

    vk_msg_id = await get_vk_by_tg(message.message_id)
    if vk_msg_id:
        tg_user_id = message.from_user.id
        name = TG_NAME_MAP.get(tg_user_id, message.from_user.full_name)

        text_content = message.text or message.caption or ""
        if not text_content:
            text_content = "[Изменено на медиа-файл]"

        new_vk_text = f"✈️ [TG] {name}:\n{text_content} ✏️"
        try:
            old_attachments = await vk_bridge.get_attachments_str(vk_msg_id)
            await vk_bridge.edit_message(
                vk_msg_id, target_vk_peer, new_vk_text, old_attachments
            )
        except Exception as e:
            logging.error(f"Ошибка редактирования в VK: {e}")


async def main():
    logging.info("🚀 Запуск StudyManager Telegram Bot...")
    await get_db_pool()
    logging.info("✅ Подключение к PostgreSQL установлено")

    await bot.delete_webhook(drop_pending_updates=True)
    task_tg = asyncio.create_task(dp.start_polling(bot))
    task_vk = asyncio.create_task(vk_bridge.start_longpoll())
    try:
        await asyncio.gather(task_tg, task_vk)
    finally:
        global db_pool
        if db_pool:
            await db_pool.close()


if __name__ == "__main__":
    asyncio.run(main())